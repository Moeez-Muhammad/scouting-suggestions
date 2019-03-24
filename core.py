from openpyxl import load_workbook
import os
import re

path = os.path.dirname(os.path.abspath(__file__))

for filename in os.listdir(path):
    if re.match("scouting.*\.xlsm", filename):
        database_name = filename

wb = load_workbook(filename=os.path.join(path, database_name), data_only=True)
ws = wb["Abbreviated NMA"]

def team(row):
    row = str(row)
    row_data = ws[row]
    team = {
        "name": ws["A" + row],
        "data": {
            "max_sandstorm_hatch": ws["M" + row], 
            "average_sandstorm_hatch": ws["N" + row], 
            "max_sandstorm_cargo": ws["O" + row], 
            "average_sandstorm_cargo": ws["P" + row],
            "max_piece": ws["C" + row],
            "max_hatch": ws["D" + row],
            "cargo_ship_rocket1_hatch": ws["E" + row],
            "higher_level_rocket_hatch": ws["F" + row],
            "max_cargo": ws["G" + row],
            "cargo_ship_rocket1_cargo": ws["H" + row],
            "higher_level_rocket_cargo": ws["I" + row],
            "climb": ws["J" + row],
            "penalty": ws["L" + row]
        },
        "info": [

        ]
    }
    for k, v in team["data"].items():
        if team["data"][k].value == "#N/A":
            team["data"][k].value = 0
    team["data"]["higher_level_rocket_average"] = (team["data"]["higher_level_rocket_hatch"].value + team["data"]["higher_level_rocket_cargo"].value)/2

    return team

red_team1 = team(4)
red_team2 = team(5)
red_team3 = team(6)

blue_team1 = team(10)
blue_team2 = team(11)
blue_team3 = team(12)