from openpyxl import load_workbook
import pdb

wb = load_workbook(filename="scouting.xlsx", data_only=True)
ws = wb["Abbreviated NMA"]

def summary(row):
    row = str(row)
    row_data = ws[row]
    print(ws["L" + row].value)
    print(ws["I" + row].value)
    team = {
        "name": ws["A" + row],
        "data": {
            "max_sandstorm_hatch": ws["M" + row], 
            "average_sandstorm_hatch": ws["N" + row], 
            "max_sandstorm_cargo": ws["O" + row], 
            "average_sandstorm_cargo": ws["P" + row],
            "max_piece": ws["C" + row],
            "max_hatch": ws["D" + row],
            "cargo_ship_rocket1_hatch": ws["E" + row],
            "higher_level_rocket_hatch": ws["F" + row],
            "max_cargo": ws["G" + row],
            "cargo_ship_rocket1_cargo": ws["H" + row],
            "higher_level_rocket_cargo": ws["I" + row],
            "higher_level_rocket_average": (ws["F" + row].value + ws["I" + row].value)/2,
            "climb": ws["J" + row],
            "penalty": ws["L" + row]
        },
        "info": [

        ]
    }

    team_name = team["name"]
    data = team["data"]
    info = team["info"]

    if data["average_sandstorm_hatch"].value - data["average_sandstorm_cargo"].value >= 0.25:
        if data["average_sandstorm_hatch"].value == 0:
            pass
        elif data["average_sandstorm_hatch"].value <= 0.5:
            info.append("Can sometimes do one sandstorm hatch")
        elif data["average_sandstorm_hatch"].value <= 0.9:
            info.append("Can mostly do one sandstorm hatch")
        elif data["average_sandstorm_hatch"].value <= 1.0:
            info.append("Can always do one sandstorm hatch")
    elif data["average_sandstorm_hatch"].value - data["average_sandstorm_cargo"].value <= -0.25:
        if data["average_sandstorm_cargo"].value == 0:
            pass
        elif data["average_sandstorm_cargo"].value <= 0.5:
            info.append("Can sometimes do one sandstorm cargo")
        elif data["average_sandstorm_cargo"].value <= 0.9:
            info.append("Can mostly do one sandstorm cargo")
        elif data["average_sandstorm_cargo"].value <= 1.0:
            info.append("Can always do one sandstorm cargo")
    else:
        if data["average_sandstorm_cargo"].value == 0:
            pass
        elif data["average_sandstorm_cargo"].value <= 0.5:
            info.append("Can sometimes do one sandstorm cargo")
        elif data["average_sandstorm_cargo"].value <= 0.9:
            info.append("Can mostly do one sandstorm cargo")
        elif data["average_sandstorm_cargo"].value <= 1.0:
            info.append("Can always do one sandstorm cargo")
        if data["average_sandstorm_hatch"].value == 0:
            pass
        elif data["average_sandstorm_hatch"].value <= 0.5:
            info.append("Can sometimes do one sandstorm hatch")
        elif data["average_sandstorm_hatch"].value <= 0.9:
            info.append("Can mostly do one sandstorm hatch")
        elif data["average_sandstorm_hatch"].value <= 1.0:
            info.append("Can always do one sandstorm hatch")
    if data["average_sandstorm_cargo"].value == 0 and data["average_sandstorm_hatch"].value == 0:
            info.append("Cannot do sandstorm")
            
    if data["max_piece"].value >= 4:
        info.append("Good gamepiece bot")
        if data["max_hatch"].value - data["max_cargo"].value > 1:
            info.append("Hatch bot")
        elif data["max_hatch"].value - data["max_cargo"].value < -1:
            info.append("Cargo bot")
        else:
            info.append("Can do both")
        if data["higher_level_rocket_average"] == 0:
            info.append("Cannot do higher level rocket")
        elif data["higher_level_rocket_average"] <= 1:
            info.append("Can somewhat do higher level rocket")
        else:
            info.append("Can do higher level rocket")
    else:
        info.append("Not very good gamepiece bot")

    if data["climb"].value < 0.85:
        info.append("Can't climb to level 1 consistently")
    elif data["climb"].value <= 1:
        info.append("Consistent level 1 climb")
    elif data["climb"].value <= 1.65:
        info.append("Sketchy level 2 climb")
    elif data["climb"].value <= 2:
        info.append("Consistent level 2 climb")
    elif data["climb"].value <= 2.65:
        info.append("Sketchy level 3 climb")
    else:
        info.append("Consistent level 3 climb")

    if data["penalty"].value < 1:
        info.append("Not very penalty prone")
    elif data["penalty"].value < 2:
        info.append("Be wary of penalties")
    else:
        info.append("VERY prone to penalties (maybe give copy of manual?)")

    team["info"] = info
    return team

red_team1 = summary(4)
red_team2 = summary(5)
red_team3 = summary(6)

blue_team1 = summary(10)
blue_team2 = summary(11)
blue_team3 = summary(12)

print(red_team1["name"].value)
for info in red_team1["info"]:
    print(info)
print(red_team2["name"].value)
for info in red_team2["info"]:
    print(info)
print(red_team3["name"].value)
for info in red_team3["info"]:
    print(info)
print(blue_team1["name"].value)
for info in blue_team1["info"]:
    print(info)
print(blue_team2["name"].value)
for info in blue_team2["info"]:
    print(info)
print(blue_team3["name"].value)
for info in blue_team3["info"]:
    print(info)